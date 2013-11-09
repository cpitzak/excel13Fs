import pymongo, sys, ystockquote

from openpyxl import Workbook
from pymongo import MongoClient

class Export13F:
  
  GURU_CELL = 0
  DATE_CELL = 3
  SYMBOL_CELL = 4
  ACTION_CELL = 5
  AVERAGE_CELL = 6
  MINIMUM_CELL = 7
  VOLUME_CELL = 8
  IMPACT_START_CELL = 9
  CURRENT_SHARE_PRICE_START_CELL = 10
  
  def __init__ (self):
    self.client = MongoClient()
    self.client = MongoClient('localhost', 27017)
    return
  
  def export(self, xlsx_file, s_flag):
    wb = Workbook()
    ws = wb.create_sheet(0)
    ws.title = "Guru 13Fs"
    print "Writing {0}...".format(xlsx_file) 
    headers = ["Gurus", "Date", "Symbol", "Action", "Average", "Minimum", "Volume", "Portfolio Impact"]
    if s_flag:
      headers.append("Current Share Price")
    
    col = 0
    for header in headers:
      header_cell = ws.cell(row = 0, column = col)
      header_cell.value = header
      if header == "Gurus":
        ws.merge_cells(start_row = 0, start_column = 0, end_row = 0, end_column = 2)
        col += 2
      col += 1
      
    db = self.client.gurudb
    gurus = db.gurus
    cursor = gurus.find()
    
    share_prices = dict()
    guruRow = 1
    for guru in cursor:
      tickers = guru['tickers']
      for ticker in tickers:
        history = ticker['transactionHistory']
        transactions = history['transactions']
        symbol = ticker['ticker']
        if share_prices.get(symbol) == None and s_flag:
          share_prices[symbol] = ystockquote.get_price(symbol)
          print "Retrieved {0} at ${1} for {2}".format(symbol, share_prices[symbol], guru['name'])
        for transaction in transactions:
          guru_cell = ws.cell(row = guruRow, column = 0)
          guru_cell.value = guru['name']
          ws.merge_cells(start_row = guruRow, start_column = 0, end_row = guruRow, end_column = 2)
          symbol_cell = ws.cell(row = guruRow, column = self.SYMBOL_CELL)
          symbol_cell.value = symbol
          impact_cell = ws.cell(row = guruRow, column = self.IMPACT_START_CELL)
          impact_cell.value = "{0}%".format(ticker['percent'])
          transaction_cell = ws.cell(row = guruRow, column = self.DATE_CELL)
          transaction_cell.value = transaction['date']
          action_cell = ws.cell(row = guruRow, column = self.ACTION_CELL)
          action_cell.value = transaction['entryType']
          average_cell = ws.cell(row = guruRow, column = self.AVERAGE_CELL)
          # convert from int to decimal. stored in mongo this way for percision reasons
          average_cell.value = float(transaction['avg']) * 0.01
          minimum_cell = ws.cell(row = guruRow, column = self.MINIMUM_CELL)
          minimum_cell.value = float(transaction['min']) * 0.01
          volume_cell = ws.cell(row = guruRow, column = self.VOLUME_CELL)
          volume_cell.value = transaction['numberOfShares']
          if s_flag:
            current_share_price_cell = ws.cell(row = guruRow, column = self.CURRENT_SHARE_PRICE_START_CELL)
            current_share_price_cell.value = share_prices.get(symbol)
          guruRow += 1
    
    wb.save(xlsx_file)
    print "Wrote {0}".format(xlsx_file) 
    return

def main():
  export13F = Export13F()
  USAGE_MESSAGE = """
  usage: export13F [-s] <file>
     
     -s  write current share price (must be online)
   
  """
  MIN_ARGS = 2
  MAX_ARGS = 3
  if len(sys.argv) < MIN_ARGS:
    sys.exit(USAGE_MESSAGE)
  
  if sys.argv[1][0] != "-":
    export13F.export(sys.argv[1], False)
  elif sys.argv[1] == "-s" and len(sys.argv) == MAX_ARGS:
    export13F.export(sys.argv[2], True)
  else:
    sys.exit(USAGE_MESSAGE)
  return
      
if __name__ == "__main__":
  main()