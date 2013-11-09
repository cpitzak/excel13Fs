import pymongo

from openpyxl import Workbook
from pymongo import MongoClient

class Export13F:
  
  def __init__ (self):
    self.client = MongoClient()
    self.client = MongoClient('localhost', 27017)
    return
  
  def export(self, xlsx_file):
    wb = Workbook()
    ws = wb.create_sheet(0)
    ws.title = "Guru 13Fs"
    
    headers = ["Gurus", "Date", "Symbol", "Action", "Average", "Minimum", "Volume", "% Change", "Impact"]
    
    for i, header in enumerate(headers):
      header_cell = ws.cell(row = 0, column = i)
      header_cell.value = header
      
    db = self.client.gurudb
    gurus = db.gurus
    cursor = gurus.find()
    
    guruRow = 0
    for guru in cursor:
      guru_cell = ws.cell(row = guruRow, column = 0)
      guru_cell.value = guru['name']
      ws.merge_cells(start_row = guruRow, start_column = 0, end_row = guruRow, end_column = 2)
      guruRow += 1
    
    wb.save(xlsx_file)
    print "Wrote {0}".format(xlsx_file) 
    return

def main():
  export13F = Export13F()
  export13F.export("gurus.xlsx")
  return
      
if __name__ == "__main__":
  main()